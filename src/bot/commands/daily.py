"""
Módulo de comandos de atualizações diárias para o Team Analysis Discord Bot.
"""

from typing import Optional
import logging
from datetime import datetime, timedelta
import os

import discord
from discord import app_commands
from discord.ext import commands
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from src.storage.feature_toggle import is_feature_enabled
from src.storage.users import get_user, get_users_by_role, check_user_is_po, get_user_display_name
from src.storage.daily import submit_daily_update, has_submitted_daily_update, get_user_daily_updates, get_all_daily_updates
from src.utils.config import get_env, get_br_time, BRAZIL_TIMEZONE, log_command, parse_date_string
from src.bot.modals import DailyUpdateModal
from src.bot.views import DailyUpdateView

logger = logging.getLogger('team_analysis_bot')


class DailyCommands(commands.Cog):
    """Comandos relacionados às atualizações diárias."""

    def __init__(self, bot: commands.Bot):
        self.bot = bot

    async def _check_daily_enabled(self, interaction: discord.Interaction) -> bool:
        """Verifica se a funcionalidade de daily está ativada."""
        if not is_feature_enabled("daily"):
            await interaction.response.send_message(
                "⚠️ A funcionalidade de atualizações diárias está temporariamente desativada.",
                ephemeral=True
            )
            log_command("INFO", interaction.user, interaction.command.name, "Funcionalidade desativada")
            return False
        return True

    @app_commands.command(name="daily", description="Envia ou atualiza sua atualização diária")
    @app_commands.describe(data="Data opcional (formatos aceitos: YYYY-MM-DD, YYYY/MM/DD ou DD/MM/YYYY)")
    async def daily_update(
        self,
        interaction: discord.Interaction,
        data: Optional[str] = None
    ):
        """
        Envia ou atualiza uma atualização diária.

        Args:
            interaction: A interação do Discord.
            data: Data opcional nos formatos YYYY-MM-DD, YYYY/MM/DD ou DD/MM/YYYY.
        """
        if not await self._check_daily_enabled(interaction):
            return

        user_id = str(interaction.user.id)
        user = get_user(user_id)

        if not user:
            await interaction.response.send_message(
                "⚠️ Você não está registrado no sistema. Peça a um administrador para registrá-lo primeiro usando o comando `/registrar`.",
                ephemeral=True
            )
            log_command("ERRO", interaction.user, "/daily", "Usuário não registrado")
            return

        daily_channel_id = get_env("DAILY_CHANNEL_ID")
        if daily_channel_id and interaction.channel_id != int(daily_channel_id):
            try:
                daily_channel = await self.bot.fetch_channel(int(daily_channel_id))
                await interaction.response.send_message(
                    f"⚠️ Por favor, use o comando `/daily` no canal {daily_channel.mention} para enviar suas atualizações diárias.",
                    ephemeral=True
                )
                log_command("ERRO", interaction.user, "/daily",
                           f"Canal incorreto. Usou no canal #{interaction.channel.name} mas deveria ser #{daily_channel.name}")
                return
            except (discord.NotFound, discord.Forbidden, ValueError):
                pass

        formatted_data = None
        if data:
            formatted_data = parse_date_string(data)
            if not formatted_data:
                await interaction.response.send_message(
                    f"⚠️ Formato de data inválido: {data}. Formatos aceitos: YYYY-MM-DD, YYYY/MM/DD ou DD/MM/YYYY.",
                    ephemeral=True
                )
                log_command("ERRO", interaction.user, f"/daily data={data}", "Formato de data inválido")
                return

            try:
                date_obj = datetime.strptime(formatted_data, "%Y-%m-%d")

                today = get_br_time().date()
                if date_obj.date() > today:
                    await interaction.response.send_message(
                        f"⚠️ Não é possível registrar atualizações para datas futuras. Hoje é {today.strftime('%d/%m/%Y')} no horário de Brasília.",
                        ephemeral=True
                    )
                    log_command("ERRO", interaction.user, f"/daily data={data}", "Data no futuro")
                    return
            except ValueError:
                await interaction.response.send_message(
                    f"⚠️ Erro ao processar a data: {data}.",
                    ephemeral=True
                )
                log_command("ERRO", interaction.user, f"/daily data={data}", "Erro ao processar data")
                return

        if has_submitted_daily_update(user_id, formatted_data):
            if formatted_data:
                date_obj = datetime.strptime(formatted_data, "%Y-%m-%d")
                date_obj = date_obj.replace(tzinfo=BRAZIL_TIMEZONE)
                formatted_date = date_obj.strftime("%d/%m/%Y")
            else:
                yesterday = get_br_time() - timedelta(days=1)
                formatted_date = yesterday.strftime("%d/%m/%Y")

            embed = discord.Embed(
                title="⚠️ Atualização Já Registrada",
                description=f"Você já enviou sua atualização diária para o dia **{formatted_date}**. Deseja atualizar este registro?",
                color=discord.Color.gold()
            )

            embed.add_field(
                name="Atenção",
                value="Se escolher atualizar, o conteúdo anterior será substituído pelo novo.",
                inline=False
            )

            command_details = f"/daily{f' data={data}' if data else ''}"
            log_command("DAILY UPDATE", interaction.user, command_details, "Usuário já enviou atualização para esta data")

            view = DailyUpdateView(interaction.user, formatted_data)
            await interaction.response.send_message(embed=embed, view=view, ephemeral=True)
            return

        command_details = f"/daily{f' data={data}' if data else ''}"
        log_command("COMANDO", interaction.user, command_details, "Iniciando modal de daily update")

        await interaction.response.send_modal(DailyUpdateModal(formatted_data, interaction.user))

    @app_commands.command(name="ver-daily", description="Visualiza suas atualizações diárias")
    @app_commands.describe(periodo="Período para visualizar (hoje, semana, mes)")
    @app_commands.choices(periodo=[
        app_commands.Choice(name="Últimos 7 dias", value="semana"),
        app_commands.Choice(name="Últimos 30 dias", value="mes")
    ])
    async def view_daily(
        self,
        interaction: discord.Interaction,
        periodo: str = "semana"
    ):
        """
        Visualiza suas atualizações diárias.

        Args:
            interaction: A interação do Discord.
            periodo: Período para visualizar.
        """
        if not await self._check_daily_enabled(interaction):
            return

        user_id = str(interaction.user.id)

        user = get_user(user_id)
        if not user:
            await interaction.response.send_message(
                "⚠️ Você não está registrado no sistema. Peça a um administrador para registrá-lo primeiro usando o comando `/registrar`.",
                ephemeral=True
            )
            log_command("ERRO", interaction.user, f"/ver-daily periodo={periodo}", "Usuário não registrado")
            return

        today = get_br_time().date()

        if periodo == "semana":
            start_date = (today - timedelta(days=7)).strftime("%Y-%m-%d")
            end_date = today.strftime("%Y-%m-%d")
            periodo_texto = "últimos 7 dias"
        elif periodo == "mes":
            start_date = (today - timedelta(days=30)).strftime("%Y-%m-%d")
            end_date = today.strftime("%Y-%m-%d")
            periodo_texto = "últimos 30 dias"

        updates = get_user_daily_updates(user_id, start_date, end_date)

        if not updates:
            await interaction.response.send_message(
                f"📝 Você não possui atualizações diárias registradas nos {periodo_texto}.",
                ephemeral=True
            )
            log_command("INFO", interaction.user, f"/ver-daily periodo={periodo}", "Nenhuma atualização encontrada")
            return

        embed = discord.Embed(
            title="📝 Suas Atualizações Diárias",
            description=f"Período: {periodo_texto}",
            color=discord.Color.blue()
        )

        for update in updates[:10]:
            report_date = datetime.strptime(update['report_date'], "%Y-%m-%d")
            report_date = report_date.replace(tzinfo=BRAZIL_TIMEZONE)
            formatted_date = report_date.strftime("%d/%m/%Y")

            content = update['content']
            if len(content) > 1024:
                content = content[:1021] + "..."

            embed.add_field(
                name=f"📅 {formatted_date}",
                value=content,
                inline=False
            )

        if len(updates) > 10:
            embed.set_footer(text=f"Mostrando 10 de {len(updates)} atualizações. Use períodos menores para ver mais detalhes. (Horário de Brasília)")
        else:
            embed.set_footer(text=f"Horário de Brasília: {get_br_time().strftime('%d/%m/%Y %H:%M:%S')}")

        await interaction.response.send_message(embed=embed, ephemeral=True)
        log_command("CONSULTA", interaction.user, f"/ver-daily periodo={periodo}",
                   f"Visualizadas {len(updates)} atualizações (mostrando {min(10, len(updates))})")

    @app_commands.command(name="relatorio-daily", description="Visualiza as atualizações diárias de todos os usuários")
    @app_commands.describe(
        data_inicial="Data inicial (formatos aceitos: YYYY-MM-DD, YYYY/MM/DD ou DD/MM/YYYY)",
        data_final="Data final (formatos aceitos: YYYY-MM-DD, YYYY/MM/DD ou DD/MM/YYYY)"
    )
    async def report_daily(
        self,
        interaction: discord.Interaction,
        data_inicial: Optional[str] = None,
        data_final: Optional[str] = None
    ):
        """
        Visualiza as atualizações diárias de todos os usuários em um período.

        Args:
            interaction: A interação do Discord.
            data_inicial: Data inicial nos formatos YYYY-MM-DD, YYYY/MM/DD ou DD/MM/YYYY (padrão: 30 dias atrás).
            data_final: Data final nos formatos YYYY-MM-DD, YYYY/MM/DD ou DD/MM/YYYY (padrão: hoje).
        """
        logger.debug(f"[DEBUG] Iniciando comando relatorio-daily: data_inicial={data_inicial}, data_final={data_final}")

        if not await self._check_daily_enabled(interaction):
            return

        admin_role_id = int(get_env("ADMIN_ROLE_ID"))
        has_permission = False

        if admin_role_id == 0:
            has_permission = interaction.user.guild_permissions.administrator
        else:
            has_permission = any(role.id == admin_role_id for role in interaction.user.roles)

        user = get_user(str(interaction.user.id))
        if user and user['role'] == 'po':
            has_permission = True

        if not has_permission:
            await interaction.response.send_message(
                "⚠️ Você não tem permissão para usar este comando. Apenas administradores e Product Owners podem ver relatórios.",
                ephemeral=True
            )
            log_command("PERMISSÃO NEGADA", interaction.user, f"/relatorio-daily data_inicial={data_inicial} data_final={data_final}")
            return

        today = get_br_time().date()

        if data_inicial:
            formatted_data_inicial = parse_date_string(data_inicial)
            if not formatted_data_inicial:
                await interaction.response.send_message(
                    f"⚠️ Formato de data inicial inválido: {data_inicial}. Formatos aceitos: YYYY-MM-DD, YYYY/MM/DD ou DD/MM/YYYY.",
                    ephemeral=True
                )
                log_command("ERRO", interaction.user, f"/relatorio-daily data_inicial={data_inicial} data_final={data_final}", "Formato de data inicial inválido")
                return
            start_date = formatted_data_inicial
        else:
            start_date = (today - timedelta(days=30)).strftime("%Y-%m-%d")

        if data_final:
            formatted_data_final = parse_date_string(data_final)
            if not formatted_data_final:
                await interaction.response.send_message(
                    f"⚠️ Formato de data final inválido: {data_final}. Formatos aceitos: YYYY-MM-DD, YYYY/MM/DD ou DD/MM/YYYY.",
                    ephemeral=True
                )
                log_command("ERRO", interaction.user, f"/relatorio-daily data_inicial={data_inicial} data_final={data_final}", "Formato de data final inválido")
                return
            end_date = formatted_data_final
        else:
            end_date = today.strftime("%Y-%m-%d")

        try:
            start_date_obj = datetime.strptime(start_date, "%Y-%m-%d").date()
            end_date_obj = datetime.strptime(end_date, "%Y-%m-%d").date()

            if start_date_obj > end_date_obj:
                await interaction.response.send_message(
                    "⚠️ A data inicial não pode ser posterior à data final.",
                    ephemeral=True
                )
                log_command("ERRO", interaction.user, f"/relatorio-daily data_inicial={data_inicial} data_final={data_final}", "Data inicial posterior à final")
                return

            if (end_date_obj - start_date_obj).days > 60:
                await interaction.response.send_message(
                    "⚠️ O período máximo para relatórios é de 60 dias.",
                    ephemeral=True
                )
                log_command("ERRO", interaction.user, f"/relatorio-daily data_inicial={data_inicial} data_final={data_final}", "Período muito longo")
                return
        except ValueError:
            await interaction.response.send_message(
                "⚠️ Formato de data inválido. Use o formato YYYY-MM-DD.",
                ephemeral=True
            )
            log_command("ERRO", interaction.user, f"/relatorio-daily data_inicial={data_inicial} data_final={data_final}", "Formato de data inválido")
            return

        await interaction.response.defer(ephemeral=True)
        log_command("PROCESSANDO", interaction.user, f"/relatorio-daily data_inicial={data_inicial} data_final={data_final}",
                   "Iniciando geração do relatório")

        logger.debug(f"[DEBUG] Buscando atualizações diárias no banco de dados...")
        all_updates = get_all_daily_updates(start_date, end_date)
        logger.debug(f"[DEBUG] Quantidade de usuários com updates: {len(all_updates)}")

        if not all_updates:
            await interaction.followup.send(
                f"📝 Não há atualizações diárias registradas no período de {start_date} a {end_date}.",
                ephemeral=True
            )
            log_command("INFO", interaction.user, f"/relatorio-daily data_inicial={data_inicial} data_final={data_final}",
                       "Nenhuma atualização encontrada")
            return

        logger.debug(f"[DEBUG] Obtendo lista de todos os usuários")
        all_users = {}
        for role in ["teammember", "po"]:
            users = get_users_by_role(role)
            for user in users:
                all_users[user["user_id"]] = {
                    "role": role,
                    "name": get_user_display_name(user["user_id"], user),
                    "user_obj": user
                }

        unique_user_ids = set()
        for user_id in all_updates.keys():
            unique_user_ids.add(user_id)

        discord_users = {}
        logger.debug(f"[DEBUG] Pré-buscando {len(unique_user_ids)} usuários do Discord em lote")
        for user_id in unique_user_ids:
            try:
                discord_user = await self.bot.fetch_user(int(user_id))
                discord_users[user_id] = discord_user
            except Exception as e:
                logger.warning(f"[DEBUG] Não foi possível buscar usuário Discord {user_id}: {str(e)}")
                discord_users[user_id] = None

        logger.debug(f"[DEBUG] Organizando atualizações para o relatório")
        sorted_updates = []
        try:
            for user_id, updates in all_updates.items():
                for update in updates:
                    sorted_updates.append({
                        'user_id': user_id,
                        'update': update
                    })

            sorted_updates.sort(key=lambda x: x['update']['report_date'], reverse=True)
        except Exception as e:
            logger.error(f"[DEBUG] Erro ao processar atualizações: {str(e)}")
            await interaction.followup.send(
                f"❌ Erro ao processar dados para o relatório: {str(e)}",
                ephemeral=True
            )
            log_command("ERRO", interaction.user, f"/relatorio-daily data_inicial={data_inicial} data_final={data_final}",
                      f"Erro ao processar dados: {str(e)}")
            return

        logger.debug(f"[DEBUG] Iniciando criação do workbook Excel...")
        wb = Workbook()
        ws = wb.active
        ws.title = "Relatório Daily"

        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        subheader_font = Font(bold=True, color="000000")
        subheader_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
        alt_row_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
        border_all = Border(top=Side(style='thin'), left=Side(style='thin'),
                           right=Side(style='thin'), bottom=Side(style='thin'))

        align_center = Alignment(horizontal="center", vertical="center")
        align_left = Alignment(horizontal="left")
        align_wrap = Alignment(wrap_text=True, vertical="top")

        headers = ["Data", "Usuário", "Papel", "Atualização", "Enviado em"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col)
            cell.value = header
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = align_center
            cell.border = border_all

        column_widths = [15, 20, 15, 60, 18]
        for i, width in enumerate(column_widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = width

        ws.row_dimensions[1].height = 25

        role_display = {
            "teammember": "Team Member",
            "po": "Product Owner"
        }

        row = 2
        logger.debug(f"[DEBUG] Preenchendo planilha com {len(sorted_updates)} atualizações")

        date_format_cache = {}
        date_obj_cache = {}

        even_rows = []
        odd_rows = []
        all_cells = []

        for idx, item in enumerate(sorted_updates):
            try:
                user_id = item['user_id']
                update = item['update']

                report_date = update['report_date']
                if report_date not in date_obj_cache:
                    date_obj = datetime.strptime(report_date, "%Y-%m-%d")
                    date_obj = date_obj.replace(tzinfo=BRAZIL_TIMEZONE)
                    date_format_cache[report_date] = date_obj.strftime("%d/%m/%Y")
                    date_obj_cache[report_date] = date_obj.date()
                formatted_date = date_format_cache[report_date]
                date_obj_value = date_obj_cache[report_date]

                is_even_row = (idx % 2 == 0)
                if is_even_row:
                    even_rows.append(row)
                else:
                    odd_rows.append(row)

                discord_user = discord_users.get(user_id)
                user_data = all_users.get(user_id, {})
                stored_user_obj = user_data.get("user_obj")

                if discord_user:
                    discord_name = discord_user.display_name
                    if stored_user_obj:
                        user_name = get_user_display_name(user_id, stored_user_obj)
                        if user_name != discord_name:
                            user_name = f"{discord_name} ({user_name})"
                    else:
                        user_name = discord_name
                else:
                    user_name = all_users.get(user_id, {}).get("name", f"Usuário {user_id}")

                user_role = all_users.get(user_id, {}).get("role", "")
                role_name = role_display.get(user_role, user_role)

                submitted_at = datetime.fromisoformat(update['submitted_at'].replace('Z', '+00:00'))
                submitted_at = submitted_at.astimezone(BRAZIL_TIMEZONE)
                formatted_submit_time = submitted_at.strftime("%d/%m/%Y %H:%M")

                submitted_at_no_tz = submitted_at.replace(tzinfo=None)

                ws.cell(row=row, column=1, value=date_obj_value)
                ws.cell(row=row, column=2, value=user_name)
                ws.cell(row=row, column=3, value=role_name)
                ws.cell(row=row, column=4, value=update['content'])
                ws.cell(row=row, column=5, value=submitted_at_no_tz)

                for col in range(1, 6):
                    all_cells.append(ws.cell(row=row, column=col))

                row += 1
            except Exception as e:
                logger.error(f"[DEBUG] Erro ao processar linha {row-1}: {str(e)}")

        logger.debug(f"[DEBUG] Aplicando formatações em lote")

        for cell in all_cells:
            cell.border = border_all

        for r in range(2, row):
            date_cell = ws.cell(row=r, column=1)
            date_cell.alignment = align_center
            date_cell.number_format = "DD/MM/YYYY"

            time_cell = ws.cell(row=r, column=5)
            time_cell.alignment = align_center
            time_cell.number_format = "DD/MM/YYYY HH:MM"
    
            ws.cell(row=r, column=2).alignment = align_left
            ws.cell(row=r, column=3).alignment = align_center
            ws.cell(row=r, column=4).alignment = align_wrap

        for r in odd_rows:
            for c in range(1, 6):
                ws.cell(row=r, column=c).fill = alt_row_fill

        logger.debug(f"[DEBUG] Finalizando formatação da planilha")
        try:
            ws.auto_filter.ref = f"A1:E{row-1}"
            ws.freeze_panes = 'A2'

            summary_row = row + 2
            ws.cell(row=summary_row, column=1, value="Resumo do Relatório").font = Font(bold=True, size=12)
            ws.merge_cells(start_row=summary_row, start_column=1, end_row=summary_row, end_column=5)
            summary_row += 1

            summary_headers = ["Estatísticas", "Valor"]
            for col, header in enumerate(summary_headers, 1):
                cell = ws.cell(row=summary_row, column=col)
                cell.value = header
                cell.font = subheader_font
                cell.fill = subheader_fill
                cell.border = border_all
                cell.alignment = align_center
            summary_row += 1

            unique_users = set(item['user_id'] for item in sorted_updates)

            summary_data = [
                ["Período do relatório", f"{start_date} a {end_date}"],
                ["Total de atualizações", len(sorted_updates)],
                ["Total de usuários", len(unique_users)],
                ["Média de atualizações por usuário", f"{len(sorted_updates)/len(unique_users):.2f}" if unique_users else "0"]
            ]

            for item in summary_data:
                ws.cell(row=summary_row, column=1, value=item[0]).border = border_all
                ws.cell(row=summary_row, column=2, value=item[1]).border = border_all
                summary_row += 1

        except Exception as e:
            logger.error(f"[DEBUG] Erro ao formatar planilha: {str(e)}")

        file_name = f"relatorio_daily_{start_date}_{end_date}.xlsx"

        try:
            logger.debug(f"[DEBUG] Salvando planilha em {file_name}")
            wb.save(file_name)
            logger.debug(f"[DEBUG] Planilha salva com sucesso")
        except Exception as e:
            logger.error(f"[DEBUG] Erro ao salvar planilha: {str(e)}")
            await interaction.followup.send(
                content=f"❌ Erro ao gerar o arquivo Excel: {str(e)}",
                ephemeral=True
            )
            log_command("ERRO", interaction.user, f"/relatorio-daily data_inicial={data_inicial} data_final={data_final}",
                       f"Erro ao salvar arquivo Excel: {str(e)}")
            return

        try:
            logger.debug(f"[DEBUG] Enviando arquivo {file_name} para o Discord")
            await interaction.followup.send(
                content=f"📊 Relatório de atualizações diárias ({start_date} a {end_date})",
                file=discord.File(file_name),
                ephemeral=True
            )
            logger.debug(f"[DEBUG] Arquivo enviado com sucesso")

            log_command("RELATÓRIO", interaction.user, f"/relatorio-daily data_inicial={data_inicial} data_final={data_final}",
                       f"Relatório Excel gerado com sucesso")

        except Exception as e:
            logger.error(f"[DEBUG] Erro ao enviar arquivo: {str(e)}")
            await interaction.followup.send(
                content=f"❌ Erro ao enviar o arquivo: {str(e)}",
                ephemeral=True
            )
            log_command("ERRO", interaction.user, f"/relatorio-daily data_inicial={data_inicial} data_final={data_final}",
                       f"Erro ao enviar arquivo: {str(e)}")

        finally:
            try:
                if os.path.exists(file_name):
                    logger.debug(f"[DEBUG] Removendo arquivo temporário {file_name}")
                    os.remove(file_name)
            except Exception as e:
                logger.error(f"[DEBUG] Erro ao remover arquivo temporário: {str(e)}")