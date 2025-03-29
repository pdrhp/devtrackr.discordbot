from typing import Optional, List, Tuple, Dict, Any
from datetime import datetime, timedelta
import os
import logging

import discord
from discord import app_commands, ui
from discord.ext import commands
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from src.storage.feature_toggle import toggle_feature, is_feature_enabled
from src.storage.time_tracking import clock_in, clock_out, get_user_records, get_all_users_records
from src.storage.users import register_user, get_user, get_users_by_role, remove_user
from src.storage.daily import submit_daily_update, get_user_daily_updates, get_all_daily_updates, get_missing_updates, clear_all_daily_updates, has_submitted_daily_update
from src.utils.config import get_env, get_br_time, to_br_timezone, BRAZIL_TIMEZONE, log_command, TIME_TRACKING_CHANNEL_ID
from src.storage.ignored_dates import add_ignored_date, get_all_ignored_dates, remove_ignored_date, parse_date_config, should_ignore_date

command_logger = logging.getLogger('team_analysis_commands')
command_logger.setLevel(logging.DEBUG)


class DateConfigModal(ui.Modal, title="Configurar Datas Ignoradas - Cobrança Daily"):
    """Modal para configurar datas ignoradas para cobrança de daily."""

    dates_config = ui.TextInput(
        label="Datas a ignorar",
        style=discord.TextStyle.paragraph,
        placeholder="Formatos: 2023-12-25 (data única), 2023-12-24-2024-01-03 (intervalo)",
        required=False
    )

    def __init__(self, bot):
        super().__init__(timeout=None)
        self.bot = bot

    async def on_submit(self, interaction: discord.Interaction):
        """Processa o envio do modal."""
        await interaction.response.defer(ephemeral=True, thinking=True)

        try:
            date_pairs = parse_date_config(self.dates_config.value)

            if not date_pairs:
                await interaction.followup.send(
                    "⚠️ Nenhuma data válida foi configurada. Por favor, verifique o formato e tente novamente.",
                    ephemeral=True
                )
                log_command("ERRO", interaction.user, "/config daily_collection", "Formato de data inválido")
                return

            existing_dates = get_all_ignored_dates()
            for date_entry in existing_dates:
                remove_ignored_date(date_entry["id"])

            success_count = 0
            for start_date, end_date in date_pairs:
                if add_ignored_date(start_date, end_date, str(interaction.user.id)):
                    success_count += 1

            formatted_dates = []
            for start_date, end_date in date_pairs:
                if start_date == end_date:
                    date_obj = datetime.strptime(start_date, "%Y-%m-%d")
                    formatted_dates.append(f"• {date_obj.strftime('%d/%m/%Y')}")
                else:
                    start_obj = datetime.strptime(start_date, "%Y-%m-%d")
                    end_obj = datetime.strptime(end_date, "%Y-%m-%d")
                    formatted_dates.append(f"• {start_obj.strftime('%d/%m/%Y')} até {end_obj.strftime('%d/%m/%Y')}")

            embed = discord.Embed(
                title="✅ Configuração de Datas Ignoradas",
                description=f"Foram configuradas {success_count} entradas de datas para serem ignoradas na cobrança de daily.",
                color=discord.Color.green()
            )

            if formatted_dates:
                embed.add_field(
                    name="📅 Datas configuradas:",
                    value="\n".join(formatted_dates),
                    inline=False
                )

            embed.set_footer(text=f"Configurado por: {interaction.user.display_name} • {get_br_time().strftime('%d/%m/%Y %H:%M:%S')}")

            await interaction.followup.send(embed=embed, ephemeral=True)
            log_command("INFO", interaction.user, "/config daily_collection", f"Configuradas {success_count} datas ignoradas")

        except Exception as e:
            logger = logging.getLogger('team_analysis_bot')
            logger.error(f"Erro ao processar o modal de configuração: {str(e)}")
            await interaction.followup.send(
                f"❌ Ocorreu um erro ao processar a configuração: {str(e)}",
                ephemeral=True
            )
            log_command("ERRO", interaction.user, "/config daily_collection", f"Erro: {str(e)}")


class AdminCommands(commands.Cog):

    def __init__(self, bot: commands.Bot):
        self.bot = bot

    @app_commands.command(name="toggle", description="Ativa/desativa funcionalidades do bot")
    @app_commands.describe(funcionalidade="Funcionalidade para ativar/desativar")
    @app_commands.choices(funcionalidade=[
        app_commands.Choice(name="Sistema de ponto", value="ponto"),
        app_commands.Choice(name="Sistema de daily", value="daily"),
        app_commands.Choice(name="Cobrança de daily", value="daily_collection"),
    ])
    async def toggle_feature(
        self,
        interaction: discord.Interaction,
        funcionalidade: str
    ):
        """
        Ativa ou desativa uma funcionalidade.

        Args:
            interaction: A interação do Discord.
            funcionalidade: A funcionalidade para alternar.
        """
        admin_role_id = int(get_env("ADMIN_ROLE_ID"))
        has_permission = False

        if admin_role_id == 0:
            has_permission = interaction.user.guild_permissions.administrator
        else:
            has_permission = any(role.id == admin_role_id for role in interaction.user.roles)

        if not has_permission:
            await interaction.response.send_message(
                "⚠️ Você não tem permissão para usar este comando.",
                ephemeral=True
            )
            log_command("PERMISSÃO NEGADA", interaction.user, f"/toggle funcionalidade={funcionalidade}")
            return

        was_enabled = is_feature_enabled(funcionalidade)
        new_state = toggle_feature(funcionalidade)

        await interaction.response.send_message(
            f"{'✅' if new_state else '❌'} Funcionalidade **{funcionalidade}** foi {'ativada' if new_state else 'desativada'}.",
            ephemeral=False
        )

        log_command("TOGGLE", interaction.user, f"/toggle funcionalidade={funcionalidade}",
                   f"Alterado de {was_enabled} para {new_state}")

    @app_commands.command(name="folha-de-ponto", description="Mostra a folha de ponto de todos os usuários")
    @app_commands.describe(periodo="Período para visualizar (hoje, semana, mes)")
    @app_commands.choices(periodo=[
        app_commands.Choice(name="Hoje", value="hoje"),
        app_commands.Choice(name="Esta semana", value="semana"),
        app_commands.Choice(name="Este mês", value="mes")
    ])
    async def timesheet(
        self,
        interaction: discord.Interaction,
        periodo: str = "hoje"
    ):
        """Comando para visualizar a folha de ponto de todos os usuários."""
        admin_role_id = int(get_env("ADMIN_ROLE_ID"))
        has_permission = False

        if admin_role_id == 0:
            has_permission = interaction.user.guild_permissions.administrator
        else:
            has_permission = any(role.id == admin_role_id for role in interaction.user.roles)

        if not has_permission:
            await interaction.response.send_message(
                "Você não tem permissão para usar este comando. Apenas administradores podem ver a folha de ponto.",
                ephemeral=True
            )
            return

        if not is_feature_enabled("ponto"):
            await interaction.response.send_message(
                "⚠️ O sistema de ponto está desativado no momento. "
                "Você pode ativá-lo com o comando `/toggle funcionalidade=ponto`.",
                ephemeral=True
            )
            return

        today = datetime.now().replace(hour=0, minute=0, second=0, microsecond=0)

        if periodo == "hoje":
            start_date = today.strftime("%Y-%m-%d")
            end_date = today.strftime("%Y-%m-%d")
            periodo_texto = "hoje"
        elif periodo == "semana":
            start_date = (today - timedelta(days=today.weekday())).strftime("%Y-%m-%d")
            end_date = today.strftime("%Y-%m-%d")
            periodo_texto = "esta semana"
        elif periodo == "mes":
            start_date = today.replace(day=1).strftime("%Y-%m-%d")
            end_date = today.strftime("%Y-%m-%d")
            periodo_texto = "este mês"

        all_users_records = get_all_users_records(start_date, end_date)

        if not all_users_records:
            await interaction.response.send_message(
                f"Não há registros de ponto para {periodo_texto}.",
                ephemeral=True
            )
            return

        await interaction.response.defer(ephemeral=True)

        summary_embed = discord.Embed(
            title=f"📊 Resumo - Folha de Ponto - {periodo_texto.capitalize()}",
            description=f"Período: {start_date} a {end_date}",
            color=discord.Color.blue()
        )

        all_embeds = []

        for user_id, records in all_users_records.items():
            try:
                member = await interaction.guild.fetch_member(int(user_id))
                user_name = member.display_name
            except:
                user_name = f"Usuário {user_id}"

            total_seconds = 0
            active_session = False

            user_embed = discord.Embed(
                title=f"📋 Detalhamento - {user_name}",
                description=f"Período: {start_date} a {end_date}",
                color=discord.Color.green()
            )

            records_text = []
            for i, record in enumerate(records, 1):
                clock_in_time = datetime.fromisoformat(record["clock_in"])

                if record["clock_out"] is None:
                    clock_out_str = "🟢 Em andamento"
                    duration_str = "Em andamento"
                    active_session = True
                else:
                    clock_out_time = datetime.fromisoformat(record["clock_out"])
                    duration = clock_out_time - clock_in_time
                    duration_seconds = duration.total_seconds()

                    hours, remainder = divmod(int(duration_seconds), 3600)
                    minutes, _ = divmod(remainder, 60)
                    duration_str = f"{hours}h {minutes}min"

                    clock_out_str = f"<t:{int(clock_out_time.timestamp())}:t>"

                    total_seconds += duration_seconds

                day_date = clock_in_time.strftime("%d/%m/%Y")

                record_str = [
                    f"**{day_date}** | Entrada: <t:{int(clock_in_time.timestamp())}:t> • "
                    f"Saída: {clock_out_str} • "
                    f"Duração: {duration_str}"
                ]

                if record.get("observation"):
                    record_str.append(f"📝 *{record['observation']}*")

                records_text.append("\n".join(record_str))

            if records_text:
                chunks = [records_text[i:i + 10] for i in range(0, len(records_text), 10)]

                for i, chunk in enumerate(chunks):
                    field_name = "Registros" if i == 0 else f"Registros (continuação {i})"
                    user_embed.add_field(
                        name=field_name,
                        value="\n\n".join(chunk),
                        inline=False
                    )
            else:
                user_embed.add_field(
                    name="Registros",
                    value="Nenhum registro completo encontrado.",
                    inline=False
                )

            total_hours, remainder = divmod(int(total_seconds), 3600)
            total_minutes, _ = divmod(remainder, 60)

            user_embed.add_field(
                name="Total de Horas",
                value=f"**{total_hours}h {total_minutes}min**",
                inline=False
            )

            status = " 🟢" if active_session else ""
            value = f"**Total: {total_hours}h {total_minutes}min**{status}"

            summary_embed.add_field(
                name=user_name,
                value=value,
                inline=True
            )

            all_embeds.append(user_embed)

        all_embeds.insert(0, summary_embed)

        await interaction.followup.send(embeds=all_embeds, ephemeral=True)

    @app_commands.command(name="limpar-resumos", description="Limpa todos os resumos diários do banco de dados (apenas para testes)")
    async def clear_daily_updates(self, interaction: discord.Interaction):
        """
        Limpa todos os resumos diários do banco de dados.
        Este comando deve ser usado apenas para testes.

        Args:
            interaction: A interação do Discord.
        """
        admin_role_id = int(get_env("ADMIN_ROLE_ID"))
        has_permission = False

        if admin_role_id == 0:
            has_permission = interaction.user.guild_permissions.administrator
        else:
            has_permission = any(role.id == admin_role_id for role in interaction.user.roles)

        if not has_permission:
            await interaction.response.send_message(
                "⚠️ Você não tem permissão para usar este comando. Apenas administradores podem limpar os resumos diários.",
                ephemeral=True
            )
            log_command("PERMISSÃO NEGADA", interaction.user, "/limpar-resumos")
            return

        embed = discord.Embed(
            title="⚠️ Confirmação: Limpar Todos os Resumos",
            description="Esta ação irá remover **PERMANENTEMENTE** todas as atualizações diárias do banco de dados.\n\n**Esta operação não pode ser desfeita.**",
            color=discord.Color.red()
        )

        embed.add_field(
            name="Tem certeza?",
            value="Este comando deve ser usado apenas para fins de teste.",
            inline=False
        )

        log_command("INICIANDO", interaction.user, "/limpar-resumos", "Solicitação de confirmação enviada")

        class ConfirmationView(discord.ui.View):
            def __init__(self, original_user_id: int):
                super().__init__(timeout=30)
                self.original_user_id = original_user_id

            @discord.ui.button(label="Sim, limpar tudo", style=discord.ButtonStyle.danger)
            async def confirm(self, btn_interaction: discord.Interaction, button: discord.ui.Button):
                if btn_interaction.user.id != self.original_user_id:
                    await btn_interaction.response.send_message("Você não pode confirmar esta ação.", ephemeral=True)
                    return

                success, message = clear_all_daily_updates()

                if success:
                    result_embed = discord.Embed(
                        title="✅ Resumos Diários Limpos",
                        description=message,
                        color=discord.Color.green()
                    )
                    log_command("SUCESSO", btn_interaction.user, "/limpar-resumos", message)
                else:
                    result_embed = discord.Embed(
                        title="❌ Erro ao Limpar Resumos",
                        description=message,
                        color=discord.Color.red()
                    )
                    log_command("ERRO", btn_interaction.user, "/limpar-resumos", f"Erro: {message}")

                await btn_interaction.response.edit_message(content=None, embed=result_embed, view=None)
                self.stop()

            @discord.ui.button(label="Cancelar", style=discord.ButtonStyle.secondary)
            async def cancel(self, btn_interaction: discord.Interaction, button: discord.ui.Button):
                if btn_interaction.user.id != self.original_user_id:
                    await btn_interaction.response.send_message("Você não pode cancelar esta ação.", ephemeral=True)
                    return

                cancel_embed = discord.Embed(
                    title="Operação Cancelada",
                    description="Nenhuma alteração foi feita.",
                    color=discord.Color.blue()
                )

                log_command("CANCELADO", btn_interaction.user, "/limpar-resumos", "Operação cancelada pelo usuário")

                await btn_interaction.response.edit_message(content=None, embed=cancel_embed, view=None)
                self.stop()

        confirmation_view = ConfirmationView(interaction.user.id)
        await interaction.response.send_message(embed=embed, view=confirmation_view, ephemeral=True)

    @app_commands.command(name="registrar", description="Registra um usuário no sistema")
    @app_commands.describe(
        tipo="Tipo de usuário (teammember: membro do time, po: product owner)",
        usuario="Usuário para registrar"
    )
    @app_commands.choices(tipo=[
        app_commands.Choice(name="Membro do time", value="teammember"),
        app_commands.Choice(name="Product Owner", value="po")
    ])
    async def register_user(
        self,
        interaction: discord.Interaction,
        tipo: str,
        usuario: discord.User
    ):
        """
        Registra um usuário no sistema.

        Args:
            interaction: A interação do Discord.
            tipo: Tipo de usuário.
            usuario: Usuário a ser registrado.
        """
        admin_role_id = int(get_env("ADMIN_ROLE_ID"))
        has_permission = False

        if admin_role_id == 0:
            has_permission = interaction.user.guild_permissions.administrator
        else:
            has_permission = any(role.id == admin_role_id for role in interaction.user.roles)

        if not has_permission:
            await interaction.response.send_message(
                "⚠️ Você não tem permissão para usar este comando.",
                ephemeral=True
            )
            log_command("PERMISSÃO NEGADA", interaction.user, f"/registrar tipo={tipo} usuario={usuario.name}")
            return

        success, message = register_user(str(usuario.id), tipo)

        if success:
            await interaction.response.send_message(
                f"✅ {message}",
                ephemeral=False
            )
            log_command("REGISTRO", interaction.user, f"/registrar tipo={tipo} usuario={usuario.name}",
                       f"Usuário registrado com sucesso como {tipo}")
        else:
            await interaction.response.send_message(
                f"⚠️ {message}",
                ephemeral=True
            )
            log_command("ERRO", interaction.user, f"/registrar tipo={tipo} usuario={usuario.name}",
                       f"Erro: {message}")

    @app_commands.command(name="remover", description="Remove um usuário do sistema")
    @app_commands.describe(usuario="Usuário para remover")
    async def remove_user(
        self,
        interaction: discord.Interaction,
        usuario: discord.User
    ):
        """
        Remove um usuário do sistema.

        Args:
            interaction: A interação do Discord.
            usuario: Usuário a ser removido.
        """
        admin_role_id = int(get_env("ADMIN_ROLE_ID"))
        has_permission = False

        if admin_role_id == 0:
            has_permission = interaction.user.guild_permissions.administrator
        else:
            has_permission = any(role.id == admin_role_id for role in interaction.user.roles)

        if not has_permission:
            await interaction.response.send_message(
                "⚠️ Você não tem permissão para usar este comando.",
                ephemeral=True
            )
            log_command("PERMISSÃO NEGADA", interaction.user, f"/remover usuario={usuario.name}")
            return

        success, message = remove_user(str(usuario.id))

        if success:
            await interaction.response.send_message(
                f"✅ {message}",
                ephemeral=False
            )
            log_command("REMOÇÃO", interaction.user, f"/remover usuario={usuario.name}",
                       "Usuário removido com sucesso")
        else:
            await interaction.response.send_message(
                f"⚠️ {message}",
                ephemeral=True
            )
            log_command("ERRO", interaction.user, f"/remover usuario={usuario.name}",
                       f"Erro: {message}")

    @app_commands.command(name="config", description="Configura opções do bot")
    @app_commands.describe(
        funcionalidade="Funcionalidade a ser configurada"
    )
    @app_commands.choices(funcionalidade=[
        app_commands.Choice(name="Cobrança de Daily", value="daily_collection")
    ])
    async def config(
        self,
        interaction: discord.Interaction,
        funcionalidade: str
    ):
        """
        Configura opções do bot.

        Args:
            interaction: A interação do Discord.
            funcionalidade: A funcionalidade a ser configurada.
        """
        admin_role_id = int(get_env("ADMIN_ROLE_ID"))
        has_permission = False

        if admin_role_id == 0:
            has_permission = interaction.user.guild_permissions.administrator
        else:
            has_permission = any(role.id == admin_role_id for role in interaction.user.roles)

        user = get_user(str(interaction.user.id))
        if user and user['role'] == 'po':
            has_permission = True

        if not has_permission:
            await interaction.response.send_message(
                "⚠️ Você não tem permissão para usar este comando. Apenas administradores e Product Owners podem configurar o bot.",
                ephemeral=True
            )
            log_command("PERMISSÃO NEGADA", interaction.user, f"/config funcionalidade={funcionalidade}")
            return

        if funcionalidade == "daily_collection":
            if not is_feature_enabled("daily"):
                await interaction.response.send_message(
                    "⚠️ A funcionalidade de daily está desativada. "
                    "Você precisa ativá-la primeiro com o comando `/toggle funcionalidade=daily`.",
                    ephemeral=True
                )
                log_command("ERRO", interaction.user, f"/config funcionalidade={funcionalidade}", "Funcionalidade de daily desativada")
                return

            if not is_feature_enabled("daily_collection"):
                await interaction.response.send_message(
                    "⚠️ A funcionalidade de cobrança de daily está desativada. "
                    "Você precisa ativá-la primeiro com o comando `/toggle funcionalidade=daily_collection`.",
                    ephemeral=True
                )
                log_command("ERRO", interaction.user, f"/config funcionalidade={funcionalidade}", "Funcionalidade de cobrança de daily desativada")
                return

            embed = discord.Embed(
                title="⚙️ Configurações - Cobrança de Daily",
                description="Escolha uma das opções abaixo para configurar:",
                color=discord.Color.blue()
            )

            embed.add_field(
                name="📅 Datas Ignoradas",
                value="Configure quais datas devem ser ignoradas na cobrança de daily. "
                      "Útil para feriados, recessos e outros períodos sem trabalho.",
                inline=False
            )

            embed.add_field(
                name="ℹ️ Formatos de Data Aceitos",
                value="• Data única: `2023-12-25`\n"
                      "• Múltiplas datas: `2023-12-25,2023-12-26`\n"
                      "• Intervalo de datas: `2023-12-24-2024-01-03`",
                inline=False
            )

            view = ConfigView(self.bot, funcionalidade)
            await interaction.response.send_message(embed=embed, view=view, ephemeral=True)
            log_command("INFO", interaction.user, f"/config funcionalidade={funcionalidade}", "Menu de opções de configuração exibido")
        else:
            await interaction.response.send_message(
                f"⚠️ A funcionalidade '{funcionalidade}' não possui opções de configuração ainda.",
                ephemeral=True
            )
            log_command("ERRO", interaction.user, f"/config funcionalidade={funcionalidade}", "Funcionalidade sem opções de configuração")

    @app_commands.command(name="remover-data-ignorada", description="Remove uma configuração de data ignorada na cobrança de daily")
    @app_commands.describe(
        id="ID da configuração de data a ser removida"
    )
    async def remove_ignored_date(
        self,
        interaction: discord.Interaction,
        id: int
    ):
        """
        Remove uma configuração de data ignorada na cobrança de daily.

        Args:
            interaction: A interação do Discord.
            id: ID da configuração a ser removida.
        """
        admin_role_id = int(get_env("ADMIN_ROLE_ID"))
        has_permission = False

        if admin_role_id == 0:
            has_permission = interaction.user.guild_permissions.administrator
        else:
            has_permission = any(role.id == admin_role_id for role in interaction.user.roles)

        user = get_user(str(interaction.user.id))
        if user and user['role'] == 'po':
            has_permission = True

        if not has_permission:
            await interaction.response.send_message(
                "⚠️ Você não tem permissão para usar este comando. Apenas administradores e Product Owners podem remover datas ignoradas.",
                ephemeral=True
            )
            log_command("PERMISSÃO NEGADA", interaction.user, f"/remover-data-ignorada id={id}")
            return

        if not is_feature_enabled("daily") or not is_feature_enabled("daily_collection"):
            await interaction.response.send_message(
                "⚠️ As funcionalidades de daily ou cobrança de daily estão desativadas.",
                ephemeral=True
            )
            log_command("ERRO", interaction.user, f"/remover-data-ignorada id={id}", "Funcionalidades desativadas")
            return

        ignored_dates = get_all_ignored_dates()
        date_to_remove = None

        for date in ignored_dates:
            if date["id"] == id:
                date_to_remove = date
                break

        if not date_to_remove:
            await interaction.response.send_message(
                f"⚠️ Não foi encontrada configuração de data ignorada com o ID {id}.",
                ephemeral=True
            )
            log_command("ERRO", interaction.user, f"/remover-data-ignorada id={id}", "ID não encontrado")
            return

        if remove_ignored_date(id):
            start_date = datetime.strptime(date_to_remove["start_date"], "%Y-%m-%d")
            end_date = datetime.strptime(date_to_remove["end_date"], "%Y-%m-%d")

            start_date_str = start_date.strftime("%d/%m/%Y")
            end_date_str = end_date.strftime("%d/%m/%Y")

            if start_date == end_date:
                date_desc = f"**{start_date_str}**"
            else:
                date_desc = f"de **{start_date_str}** até **{end_date_str}**"

            await interaction.response.send_message(
                f"✅ Configuração de data ignorada {date_desc} (ID: {id}) foi removida com sucesso.",
                ephemeral=True
            )
            log_command("INFO", interaction.user, f"/remover-data-ignorada id={id}", "Data removida com sucesso")
        else:
            await interaction.response.send_message(
                f"❌ Ocorreu um erro ao tentar remover a configuração de data ignorada com ID {id}.",
                ephemeral=True
            )
            log_command("ERRO", interaction.user, f"/remover-data-ignorada id={id}", "Erro ao remover")

    @app_commands.command(name="listar-datas-ignoradas", description="Lista as datas configuradas para serem ignoradas na cobrança de daily")
    async def list_ignored_dates(self, interaction: discord.Interaction):
        """
        Lista as datas configuradas para serem ignoradas na cobrança de daily.

        Args:
            interaction: A interação do Discord.
        """
        admin_role_id = int(get_env("ADMIN_ROLE_ID"))
        has_permission = False

        if admin_role_id == 0:
            has_permission = interaction.user.guild_permissions.administrator
        else:
            has_permission = any(role.id == admin_role_id for role in interaction.user.roles)

        user = get_user(str(interaction.user.id))
        if user and user['role'] == 'po':
            has_permission = True

        if not has_permission:
            await interaction.response.send_message(
                "⚠️ Você não tem permissão para usar este comando. Apenas administradores e Product Owners podem ver as datas ignoradas.",
                ephemeral=True
            )
            log_command("PERMISSÃO NEGADA", interaction.user, "/listar-datas-ignoradas")
            return

        if not is_feature_enabled("daily") or not is_feature_enabled("daily_collection"):
            await interaction.response.send_message(
                "⚠️ As funcionalidades de daily ou cobrança de daily estão desativadas.",
                ephemeral=True
            )
            log_command("ERRO", interaction.user, "/listar-datas-ignoradas", "Funcionalidades desativadas")
            return

        ignored_dates = get_all_ignored_dates()

        if not ignored_dates:
            await interaction.response.send_message(
                "📅 Não há datas configuradas para serem ignoradas na cobrança de daily.",
                ephemeral=True
            )
            log_command("INFO", interaction.user, "/listar-datas-ignoradas", "Nenhuma data configurada")
            return

        embed = discord.Embed(
            title="📅 Datas Ignoradas - Cobrança de Daily",
            description="Estas são as datas configuradas para serem ignoradas na cobrança de daily:",
            color=discord.Color.blue()
        )

        for date_entry in ignored_dates:
            start_date = datetime.strptime(date_entry["start_date"], "%Y-%m-%d")
            end_date = datetime.strptime(date_entry["end_date"], "%Y-%m-%d")
            created_at = datetime.strptime(date_entry["created_at"], "%Y-%m-%d %H:%M:%S")

            start_date_str = start_date.strftime("%d/%m/%Y")
            end_date_str = end_date.strftime("%d/%m/%Y")
            created_at_str = created_at.strftime("%d/%m/%Y %H:%M:%S")

            if start_date == end_date:
                date_desc = f"📆 **{start_date_str}**"
            else:
                date_desc = f"📆 De **{start_date_str}** até **{end_date_str}**"

            try:
                creator_user = await self.bot.fetch_user(int(date_entry["created_by"]))
                creator_name = creator_user.display_name
            except:
                creator_name = f"Usuário {date_entry['created_by']}"

            embed.add_field(
                name=f"ID: {date_entry['id']} - {date_desc}",
                value=f"Configurado por: {creator_name} em {created_at_str}",
                inline=False
            )

        embed.set_footer(text=f"Total: {len(ignored_dates)} configurações • ID pode ser usado com /remover-data-ignorada")

        await interaction.response.send_message(embed=embed, ephemeral=True)
        log_command("CONSULTA", interaction.user, "/listar-datas-ignoradas", f"Listadas {len(ignored_dates)} configurações")

    @app_commands.command(name="testar-datas-ignoradas", description="Testa se uma data específica está configurada para ser ignorada")
    @app_commands.describe(
        data="Data para testar no formato YYYY-MM-DD"
    )
    async def test_ignored_date(
        self,
        interaction: discord.Interaction,
        data: str
    ):
        """
        Testa se uma data específica está configurada para ser ignorada na cobrança de daily.

        Args:
            interaction: A interação do Discord.
            data: Data para testar.
        """
        admin_role_id = int(get_env("ADMIN_ROLE_ID"))
        has_permission = False

        if admin_role_id == 0:
            has_permission = interaction.user.guild_permissions.administrator
        else:
            has_permission = any(role.id == admin_role_id for role in interaction.user.roles)

        user = get_user(str(interaction.user.id))
        if user and user['role'] == 'po':
            has_permission = True

        if not has_permission:
            await interaction.response.send_message(
                "⚠️ Você não tem permissão para usar este comando.",
                ephemeral=True
            )
            log_command("PERMISSÃO NEGADA", interaction.user, f"/testar-datas-ignoradas data={data}")
            return

        try:
            date_obj = datetime.strptime(data, "%Y-%m-%d")

            is_ignored = should_ignore_date(date_obj)

            if is_ignored:
                await interaction.response.send_message(
                    f"✅ A data **{date_obj.strftime('%d/%m/%Y')}** está configurada para ser ignorada na cobrança de daily.",
                    ephemeral=True
                )
            else:
                await interaction.response.send_message(
                    f"ℹ️ A data **{date_obj.strftime('%d/%m/%Y')}** NÃO está configurada para ser ignorada na cobrança de daily.",
                    ephemeral=True
                )

            log_command("INFO", interaction.user, f"/testar-datas-ignoradas data={data}", f"Resultado: {is_ignored}")
        except ValueError:
            await interaction.response.send_message(
                f"⚠️ Formato de data inválido: {data}. Use o formato YYYY-MM-DD.",
                ephemeral=True
            )
            log_command("ERRO", interaction.user, f"/testar-datas-ignoradas data={data}", "Formato de data inválido")


class TimeTrackingCommands(commands.Cog):

    def __init__(self, bot: commands.Bot):
        self.bot = bot

    async def handle_disabled_feature(self, interaction: discord.Interaction):
        """Trata interação quando a funcionalidade de controle de horas está desativada."""
        await interaction.response.send_message(
            "⚠️ O sistema de ponto está desativado no momento. "
            "Um administrador pode ativá-lo com o comando `/toggle funcionalidade=ponto`.",
            ephemeral=True
        )
        log_command("FEATURE DESATIVADA", interaction.user, "/on ou /off", "Sistema de ponto desativado")

    async def check_channel(self, interaction: discord.Interaction):
        """Verifica se o comando está sendo usado no canal correto."""
        time_tracking_channel_id = get_env(TIME_TRACKING_CHANNEL_ID, "0")

        if time_tracking_channel_id == "0":
            return True

        if interaction.channel_id != int(time_tracking_channel_id):
            try:
                time_channel = await self.bot.fetch_channel(int(time_tracking_channel_id))
                await interaction.response.send_message(
                    f"⚠️ Por favor, use os comandos de controle de ponto no canal {time_channel.mention}.",
                    ephemeral=True
                )
                log_command("ERRO", interaction.user, f"/{interaction.command.name}",
                          f"Canal incorreto. Usou no canal #{interaction.channel.name} mas deveria ser #{time_channel.name}")
                return False
            except (discord.NotFound, discord.Forbidden, ValueError):
                return True

        return True

    @app_commands.command(name="on", description="Registra entrada e inicia contagem de horas")
    @app_commands.describe(observacao="Observação opcional sobre este registro de entrada")
    async def clock_in(self, interaction: discord.Interaction, observacao: Optional[str] = None):
        """Comando de registro de entrada."""
        if not is_feature_enabled("ponto"):
            await self.handle_disabled_feature(interaction)
            return

        if not await self.check_channel(interaction):
            return

        user_id = str(interaction.user.id)
        success, message = clock_in(user_id, observacao)

        if success:
            timestamp = discord.utils.format_dt(discord.utils.utcnow(), style='F')
            br_time = get_br_time().strftime("%d/%m/%Y %H:%M:%S")

            message_content = [
                f"🕒 **{interaction.user.display_name}** registrou ponto com sucesso!",
                f"**Entrada:** {timestamp}",
                f"**Horário de Brasília:** {br_time}"
            ]

            if observacao:
                message_content.append(f"**Observação:** {observacao}")

            await interaction.response.send_message(
                "\n".join(message_content),
                ephemeral=False
            )

            log_details = f"Entrada registrada às {br_time}"
            if observacao:
                log_details += f" | Observação: {observacao}"

            log_command("ENTRADA", interaction.user, "/on", log_details)
        else:
            await interaction.response.send_message(
                f"⚠️ {message}",
                ephemeral=True
            )

            log_command("ERRO", interaction.user, "/on", f"Erro: {message}")

    @app_commands.command(name="off", description="Registra saída e para contagem de horas")
    @app_commands.describe(observacao="Observação opcional sobre este registro de saída")
    async def clock_out(self, interaction: discord.Interaction, observacao: Optional[str] = None):
        """Comando de registro de saída."""
        if not is_feature_enabled("ponto"):
            await self.handle_disabled_feature(interaction)
            return

        if not await self.check_channel(interaction):
            return

        user_id = str(interaction.user.id)
        success, message, duration = clock_out(user_id, observacao)

        if success:
            timestamp = discord.utils.format_dt(discord.utils.utcnow(), style='F')
            br_time = get_br_time().strftime("%d/%m/%Y %H:%M:%S")

            message_content = [
                f"🕒 **{interaction.user.display_name}** registrou ponto com sucesso!",
                f"**Saída:** {timestamp}",
                f"**Horário de Brasília:** {br_time}",
                f"**Tempo trabalhado:** {duration}"
            ]

            if observacao:
                message_content.append(f"**Observação:** {observacao}")

            await interaction.response.send_message(
                "\n".join(message_content),
                ephemeral=False
            )

            log_details = f"Saída registrada às {br_time}, duração: {duration}"
            if observacao:
                log_details += f" | Observação: {observacao}"

            log_command("SAÍDA", interaction.user, "/off", log_details)
        else:
            await interaction.response.send_message(
                f"⚠️ {message}",
                ephemeral=True
            )

            log_command("ERRO", interaction.user, "/off", f"Erro: {message}")


class UserCommands(commands.Cog):
    """Comandos relacionados a gerenciamento de usuários."""

    def __init__(self, bot: commands.Bot):
        self.bot = bot

    @app_commands.command(name="listar-usuarios", description="Lista todos os usuários registrados")
    @app_commands.describe(tipo="Tipo de usuário a listar")
    @app_commands.choices(tipo=[
        app_commands.Choice(name="Team Members", value="teammember"),
        app_commands.Choice(name="Product Owners", value="po"),
        app_commands.Choice(name="Todos", value="all")
    ])
    async def list_users(
        self,
        interaction: discord.Interaction,
        tipo: str = "all"
    ):
        """
        Lista todos os usuários registrados.

        Args:
            interaction: A interação do Discord.
            tipo: Tipo de usuário a listar.
        """
        admin_role_id = int(get_env("ADMIN_ROLE_ID"))
        has_permission = False

        if admin_role_id == 0:
            has_permission = interaction.user.guild_permissions.administrator
        else:
            has_permission = any(role.id == admin_role_id for role in interaction.user.roles)

        if not has_permission:
            await interaction.response.send_message(
                "⚠️ Você não tem permissão para usar este comando.",
                ephemeral=True
            )
            log_command("PERMISSÃO NEGADA", interaction.user, f"/listar-usuarios tipo={tipo}")
            return

        users = get_users_by_role(tipo)

        if not users:
            await interaction.response.send_message(
                f"⚠️ Não há usuários do tipo '{tipo}' registrados.",
                ephemeral=True
            )
            log_command("INFO", interaction.user, f"/listar-usuarios tipo={tipo}", "Nenhum usuário encontrado")
            return

        user_strings = []
        for user_data in users:
            user_id = user_data["user_id"]

            try:
                user = await self.bot.fetch_user(int(user_id))
                display_name = user.display_name
                user_string = f"• {user.mention} ({display_name})"
            except:
                user_string = f"• ID: {user_id} (Usuário não encontrado)"

            user_strings.append(user_string)

        tipo_display = {
            "teammember": "Team Members",
            "po": "Product Owners",
            "all": "Todos os Usuários"
        }.get(tipo, tipo)

        embed = discord.Embed(
            title=f"📋 Lista de Usuários: {tipo_display}",
            description="\n".join(user_strings) if user_strings else "Nenhum usuário encontrado.",
            color=discord.Color.blue()
        )

        embed.set_footer(text=f"Total: {len(users)} usuários")

        await interaction.response.send_message(embed=embed, ephemeral=True)
        log_command("LISTAGEM", interaction.user, f"/listar-usuarios tipo={tipo}", f"Listados {len(users)} usuários")


class DailyUpdateModal(ui.Modal, title="Atualização Diária"):
    """Modal para input de atualização diária."""

    daily_content = ui.TextInput(
        label="O que você fez?",
        style=discord.TextStyle.paragraph,
        placeholder="Descreva o que você fez no dia anterior...",
        required=True,
        min_length=10,
        max_length=1000
    )

    def __init__(self, report_date: Optional[str] = None, user: discord.User = None):
        super().__init__()
        self.report_date = report_date
        self.user = user

        if report_date:
            date_obj = datetime.strptime(report_date, "%Y-%m-%d")
            date_obj = date_obj.replace(tzinfo=BRAZIL_TIMEZONE)
            date_formatted = date_obj.strftime("%d/%m/%Y")
            self.daily_content.placeholder = f"Descreva o que você fez em {date_formatted}..."

    async def on_submit(self, interaction: discord.Interaction):
        """Chamado quando o usuário envia o formulário."""
        user_id = str(interaction.user.id)
        content = self.daily_content.value

        success, message = submit_daily_update(user_id, content, self.report_date)

        if success:
            if self.report_date:
                date_obj = datetime.strptime(self.report_date, "%Y-%m-%d")
                date_obj = date_obj.replace(tzinfo=BRAZIL_TIMEZONE)
                date_formatted = date_obj.strftime("%d/%m/%Y")
                date_display = f"**{date_formatted}**"
                report_date_log = date_formatted
            else:
                yesterday = get_br_time() - timedelta(days=1)
                date_display = f"**{yesterday.strftime('%d/%m/%Y')}**"
                report_date_log = yesterday.strftime("%d/%m/%Y")

            embed = discord.Embed(
                title="📝 Nova Atualização Diária",
                description=f"**{interaction.user.display_name}** enviou uma atualização para {date_display}",
                color=discord.Color.green()
            )

            embed.add_field(
                name="Conteúdo",
                value=content,
                inline=False
            )

            current_time_br = get_br_time().strftime("%d/%m/%Y %H:%M:%S")
            embed.set_footer(text=f"Horário de Brasília: {current_time_br}")

            await interaction.response.send_message(embed=embed, ephemeral=False)

            log_command("DAILY UPDATE", interaction.user, "/daily",
                       f"Atualização para {report_date_log} registrada com sucesso ({len(content)} caracteres)")
        else:
            await interaction.response.send_message(
                f"⚠️ {message}",
                ephemeral=True
            )

            log_command("ERRO", interaction.user, "/daily", f"Erro: {message}")


class DailyUpdateView(ui.View):
    """View com botões para atualizar ou cancelar uma atualização diária existente."""

    def __init__(self, user: discord.User, report_date: Optional[str] = None):
        super().__init__(timeout=300)
        self.user = user
        self.report_date = report_date

    @ui.button(label="Atualizar", style=discord.ButtonStyle.primary, emoji="📝")
    async def update_button(self, interaction: discord.Interaction, button: ui.Button):
        """Botão para atualizar uma daily já existente."""
        if interaction.user.id != self.user.id:
            await interaction.response.send_message(
                "Você não pode interagir com estes botões, pois não são destinados a você.",
                ephemeral=True
            )
            return

        await interaction.response.send_modal(DailyUpdateModal(self.report_date, self.user))
        log_command("DAILY UPDATE", interaction.user, f"/daily{f' data={self.report_date}' if self.report_date else ''}", "Modal aberto para atualização")

    @ui.button(label="Cancelar", style=discord.ButtonStyle.secondary, emoji="❌")
    async def cancel_button(self, interaction: discord.Interaction, button: ui.Button):
        """Botão para cancelar a atualização."""
        if interaction.user.id != self.user.id:
            await interaction.response.send_message(
                "Você não pode interagir com estes botões, pois não são destinados a você.",
                ephemeral=True
            )
            return

        await interaction.response.edit_message(
            content="Atualização cancelada.",
            view=None
        )
        log_command("DAILY UPDATE", interaction.user, f"/daily{f' data={self.report_date}' if self.report_date else ''}", "Atualização cancelada pelo usuário")


class DailyCommands(commands.Cog):
    """Comandos relacionados às atualizações diárias."""

    def __init__(self, bot: commands.Bot):
        self.bot = bot

    async def _check_daily_enabled(self, interaction: discord.Interaction) -> bool:
        """Verifica se a funcionalidade de daily está ativada."""
        if not is_feature_enabled("daily"):
            await interaction.response.send_message(
                "⚠️ A funcionalidade de atualizações diárias está temporariamente desativada.",
                ephemeral=True
            )
            log_command("INFO", interaction.user, interaction.command.name, "Funcionalidade desativada")
            return False
        return True

    @app_commands.command(name="daily", description="Envia ou atualiza sua atualização diária")
    @app_commands.describe(data="Data opcional no formato YYYY-MM-DD (padrão: dia anterior)")
    async def daily_update(
        self,
        interaction: discord.Interaction,
        data: Optional[str] = None
    ):
        """
        Envia ou atualiza uma atualização diária.

        Args:
            interaction: A interação do Discord.
            data: Data opcional no formato YYYY-MM-DD.
        """
        if not await self._check_daily_enabled(interaction):
            return

        user_id = str(interaction.user.id)
        user = get_user(user_id)

        if not user:
            await interaction.response.send_message(
                "⚠️ Você não está registrado no sistema. Peça a um administrador para registrá-lo primeiro usando o comando `/registrar`.",
                ephemeral=True
            )
            log_command("ERRO", interaction.user, "/daily", "Usuário não registrado")
            return

        daily_channel_id = get_env("DAILY_CHANNEL_ID")
        if daily_channel_id and interaction.channel_id != int(daily_channel_id):
            try:
                daily_channel = await self.bot.fetch_channel(int(daily_channel_id))
                await interaction.response.send_message(
                    f"⚠️ Por favor, use o comando `/daily` no canal {daily_channel.mention} para enviar suas atualizações diárias.",
                    ephemeral=True
                )
                log_command("ERRO", interaction.user, "/daily",
                           f"Canal incorreto. Usou no canal #{interaction.channel.name} mas deveria ser #{daily_channel.name}")
                return
            except (discord.NotFound, discord.Forbidden, ValueError):
                pass

        if data:
            try:
                date_obj = datetime.strptime(data, "%Y-%m-%d")

                today = get_br_time().date()
                if date_obj.date() > today:
                    await interaction.response.send_message(
                        f"⚠️ Não é possível registrar atualizações para datas futuras. Hoje é {today.strftime('%d/%m/%Y')} no horário de Brasília.",
                        ephemeral=True
                    )
                    log_command("ERRO", interaction.user, f"/daily data={data}", "Data no futuro")
                    return
            except ValueError:
                await interaction.response.send_message(
                    f"⚠️ Formato de data inválido: {data}. Use o formato YYYY-MM-DD.",
                    ephemeral=True
                )
                log_command("ERRO", interaction.user, f"/daily data={data}", "Formato de data inválido")
                return

        if has_submitted_daily_update(user_id, data):
            if data:
                date_obj = datetime.strptime(data, "%Y-%m-%d")
                date_obj = date_obj.replace(tzinfo=BRAZIL_TIMEZONE)
                formatted_date = date_obj.strftime("%d/%m/%Y")
            else:
                yesterday = get_br_time() - timedelta(days=1)
                formatted_date = yesterday.strftime("%d/%m/%Y")

            embed = discord.Embed(
                title="⚠️ Atualização Já Registrada",
                description=f"Você já enviou sua atualização diária para o dia **{formatted_date}**. Deseja atualizar este registro?",
                color=discord.Color.gold()
            )

            embed.add_field(
                name="Atenção",
                value="Se escolher atualizar, o conteúdo anterior será substituído pelo novo.",
                inline=False
            )

            command_details = f"/daily{f' data={data}' if data else ''}"
            log_command("DAILY UPDATE", interaction.user, command_details, "Usuário já enviou atualização para esta data")

            view = DailyUpdateView(interaction.user, data)
            await interaction.response.send_message(embed=embed, view=view, ephemeral=True)
            return

        command_details = f"/daily{f' data={data}' if data else ''}"
        log_command("COMANDO", interaction.user, command_details, "Iniciando modal de daily update")

        await interaction.response.send_modal(DailyUpdateModal(data, interaction.user))

    @app_commands.command(name="ver-daily", description="Visualiza suas atualizações diárias")
    @app_commands.describe(periodo="Período para visualizar (hoje, semana, mes)")
    @app_commands.choices(periodo=[
        app_commands.Choice(name="Últimos 7 dias", value="semana"),
        app_commands.Choice(name="Últimos 30 dias", value="mes")
    ])
    async def view_daily(
        self,
        interaction: discord.Interaction,
        periodo: str = "semana"
    ):
        """
        Visualiza suas atualizações diárias.

        Args:
            interaction: A interação do Discord.
            periodo: Período para visualizar.
        """
        if not await self._check_daily_enabled(interaction):
            return

        user_id = str(interaction.user.id)

        user = get_user(user_id)
        if not user:
            await interaction.response.send_message(
                "⚠️ Você não está registrado no sistema. Peça a um administrador para registrá-lo primeiro usando o comando `/registrar`.",
                ephemeral=True
            )
            log_command("ERRO", interaction.user, f"/ver-daily periodo={periodo}", "Usuário não registrado")
            return

        today = get_br_time().date()

        if periodo == "semana":
            start_date = (today - timedelta(days=7)).strftime("%Y-%m-%d")
            end_date = today.strftime("%Y-%m-%d")
            periodo_texto = "últimos 7 dias"
        elif periodo == "mes":
            start_date = (today - timedelta(days=30)).strftime("%Y-%m-%d")
            end_date = today.strftime("%Y-%m-%d")
            periodo_texto = "últimos 30 dias"

        updates = get_user_daily_updates(user_id, start_date, end_date)

        if not updates:
            await interaction.response.send_message(
                f"📝 Você não possui atualizações diárias registradas nos {periodo_texto}.",
                ephemeral=True
            )
            log_command("INFO", interaction.user, f"/ver-daily periodo={periodo}", "Nenhuma atualização encontrada")
            return

        embed = discord.Embed(
            title="📝 Suas Atualizações Diárias",
            description=f"Período: {periodo_texto}",
            color=discord.Color.blue()
        )

        for update in updates[:10]:
            report_date = datetime.strptime(update['report_date'], "%Y-%m-%d")
            report_date = report_date.replace(tzinfo=BRAZIL_TIMEZONE)
            formatted_date = report_date.strftime("%d/%m/%Y")

            content = update['content']
            if len(content) > 1024:
                content = content[:1021] + "..."

            embed.add_field(
                name=f"📅 {formatted_date}",
                value=content,
                inline=False
            )

        if len(updates) > 10:
            embed.set_footer(text=f"Mostrando 10 de {len(updates)} atualizações. Use períodos menores para ver mais detalhes. (Horário de Brasília)")
        else:
            embed.set_footer(text=f"Horário de Brasília: {get_br_time().strftime('%d/%m/%Y %H:%M:%S')}")

        await interaction.response.send_message(embed=embed, ephemeral=True)
        log_command("CONSULTA", interaction.user, f"/ver-daily periodo={periodo}",
                   f"Visualizadas {len(updates)} atualizações (mostrando {min(10, len(updates))})")

    @app_commands.command(name="relatorio-daily", description="Visualiza as atualizações diárias de todos os usuários")
    @app_commands.describe(
        data_inicial="Data inicial no formato YYYY-MM-DD (padrão: 30 dias atrás)",
        data_final="Data final no formato YYYY-MM-DD (padrão: hoje)"
    )
    async def report_daily(
        self,
        interaction: discord.Interaction,
        data_inicial: Optional[str] = None,
        data_final: Optional[str] = None
    ):
        """
        Visualiza as atualizações diárias de todos os usuários em um período.

        Args:
            interaction: A interação do Discord.
            data_inicial: Data inicial no formato YYYY-MM-DD (padrão: 30 dias atrás).
            data_final: Data final no formato YYYY-MM-DD (padrão: hoje).
        """
        logger = logging.getLogger('team_analysis_commands')
        logger.debug(f"[DEBUG] Iniciando comando relatorio-daily: data_inicial={data_inicial}, data_final={data_final}")

        admin_role_id = int(get_env("ADMIN_ROLE_ID"))
        has_permission = False

        if admin_role_id == 0:
            has_permission = interaction.user.guild_permissions.administrator
        else:
            has_permission = any(role.id == admin_role_id for role in interaction.user.roles)

        user = get_user(str(interaction.user.id))
        if user and user['role'] == 'po':
            has_permission = True

        if not has_permission:
            await interaction.response.send_message(
                "⚠️ Você não tem permissão para usar este comando. Apenas administradores e Product Owners podem ver relatórios de atualizações diárias.",
                ephemeral=True
            )
            log_command("PERMISSÃO NEGADA", interaction.user, f"/relatorio-daily data_inicial={data_inicial} data_final={data_final}")
            return

        today = get_br_time().date()
        if data_final is None:
            data_final = today.strftime("%Y-%m-%d")

        if data_inicial is None:
            data_inicial = (today - timedelta(days=30)).strftime("%Y-%m-%d")

        logger.debug(f"[DEBUG] Datas calculadas: data_inicial={data_inicial}, data_final={data_final}")

        try:
            start_date_obj = datetime.strptime(data_inicial, "%Y-%m-%d")
            end_date_obj = datetime.strptime(data_final, "%Y-%m-%d")

            if start_date_obj > end_date_obj:
                await interaction.response.send_message(
                    "⚠️ A data inicial deve ser anterior ou igual à data final.",
                    ephemeral=True
                )
                log_command("ERRO", interaction.user, f"/relatorio-daily data_inicial={data_inicial} data_final={data_final}",
                           "Data inicial posterior à data final")
                return

        except ValueError:
            await interaction.response.send_message(
                "⚠️ Formato de data inválido. Use o formato YYYY-MM-DD.",
                ephemeral=True
            )
            log_command("ERRO", interaction.user, f"/relatorio-daily data_inicial={data_inicial} data_final={data_final}",
                       "Formato de data inválido")
            return

        await interaction.response.defer(ephemeral=True)
        log_command("PROCESSANDO", interaction.user, f"/relatorio-daily data_inicial={data_inicial} data_final={data_final}",
                   "Iniciando geração do relatório")

        logger.debug(f"[DEBUG] Buscando atualizações diárias no banco de dados...")
        all_updates = get_all_daily_updates(data_inicial, data_final)
        logger.debug(f"[DEBUG] Quantidade de usuários com updates: {len(all_updates)}")

        if not all_updates:
            await interaction.followup.send(
                f"📝 Não há atualizações diárias registradas no período de {data_inicial} a {data_final}.",
                ephemeral=True
            )
            log_command("INFO", interaction.user, f"/relatorio-daily data_inicial={data_inicial} data_final={data_final}",
                       "Nenhuma atualização encontrada")
            return

        logger.debug(f"[DEBUG] Iniciando criação do workbook Excel...")
        wb = Workbook()
        ws = wb.active
        ws.title = "Relatório Daily"

        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        subheader_font = Font(bold=True, color="000000")
        subheader_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
        date_fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
        alt_row_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
        border_bottom = Border(bottom=Side(style='thin'))
        border_all = Border(top=Side(style='thin'), left=Side(style='thin'),
                           right=Side(style='thin'), bottom=Side(style='thin'))

        headers = ["Data", "Usuário", "Papel", "Atualização", "Enviado em"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col)
            cell.value = header
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = border_all

        logger.debug(f"[DEBUG] Cabeçalhos da planilha configurados")

        role_display = {
            "teammember": "Team Member",
            "po": "Product Owner"
        }

        row = 2
        current_date = None
        use_alt_color = False

        logger.debug(f"[DEBUG] Obtendo lista de todos os usuários")
        all_users = {}
        for role in ["teammember", "po"]:
            users = get_users_by_role(role)
            for user in users:
                all_users[user["user_id"]] = {
                    "role": role,
                    "name": user["user_name"]
                }

        logger.debug(f"[DEBUG] Recuperados {len(all_users)} usuários no total")

        logger.debug(f"[DEBUG] Organizando atualizações para o relatório")
        sorted_updates = []

        try:
            for user_id, updates in all_updates.items():
                logger.debug(f"[DEBUG] Processando {len(updates)} atualizações para o usuário {user_id}")
                for update in updates:
                    sorted_updates.append({
                        'user_id': user_id,
                        'update': update
                    })

            logger.debug(f"[DEBUG] Total de atualizações coletadas: {len(sorted_updates)}")
            sorted_updates.sort(key=lambda x: x['update']['report_date'], reverse=True)
            logger.debug(f"[DEBUG] Atualizações ordenadas por data")
        except Exception as e:
            logger.error(f"[DEBUG] Erro ao processar atualizações: {str(e)}")
            await interaction.followup.send(
                f"❌ Erro ao processar dados para o relatório: {str(e)}",
                ephemeral=True
            )
            log_command("ERRO", interaction.user, f"/relatorio-daily data_inicial={data_inicial} data_final={data_final}",
                      f"Erro ao processar dados: {str(e)}")
            return

        logger.debug(f"[DEBUG] Preenchendo planilha com {len(sorted_updates)} atualizações")
        for item in sorted_updates:
            try:
                user_id = item['user_id']
                update = item['update']

                date_obj = datetime.strptime(update['report_date'], "%Y-%m-%d")
                date_obj = date_obj.replace(tzinfo=BRAZIL_TIMEZONE)
                formatted_date = date_obj.strftime("%d/%m/%Y")

                use_alt_color = not use_alt_color
                row_fill = alt_row_fill if use_alt_color else None

                try:
                    user_obj = await self.bot.fetch_user(int(user_id))
                    user_name = user_obj.display_name
                except Exception as e:
                    logger.warning(f"[DEBUG] Não foi possível buscar usuário Discord {user_id}: {str(e)}")
                    user_name = all_users.get(user_id, {}).get("name", f"Usuário {user_id}")

                user_role = all_users.get(user_id, {}).get("role", "")
                role_name = role_display.get(user_role, user_role)

                submitted_at = datetime.fromisoformat(update['submitted_at'].replace('Z', '+00:00'))
                submitted_at = submitted_at.astimezone(BRAZIL_TIMEZONE)
                formatted_submit_time = submitted_at.strftime("%d/%m/%Y %H:%M")

                ws.cell(row=row, column=1, value=formatted_date).alignment = Alignment(horizontal="center")
                ws.cell(row=row, column=2, value=user_name).alignment = Alignment(horizontal="left")
                ws.cell(row=row, column=3, value=role_name).alignment = Alignment(horizontal="center")

                content_cell = ws.cell(row=row, column=4, value=update['content'])
                content_cell.alignment = Alignment(wrap_text=True, vertical="top")

                ws.cell(row=row, column=5, value=formatted_submit_time).alignment = Alignment(horizontal="center")

                for col in range(1, 6):
                    cell = ws.cell(row=row, column=col)
                    cell.border = border_all
                    if row_fill:
                        cell.fill = row_fill

                row += 1
            except Exception as e:
                logger.error(f"[DEBUG] Erro ao processar linha {row-1}: {str(e)}")

        logger.debug(f"[DEBUG] Finalizando formatação da planilha")
        try:
            ws.auto_filter.ref = f"A1:E{row-1}"
            ws.freeze_panes = 'A2'

            column_widths = [15, 20, 15, 60, 18]
            for i, width in enumerate(column_widths, 1):
                ws.column_dimensions[get_column_letter(i)].width = width

            ws.row_dimensions[1].height = 25

            row += 2

            ws.cell(row=row, column=1, value="Resumo do Relatório").font = Font(bold=True, size=12)
            ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=5)
            row += 1

            summary_headers = ["Estatísticas", "Valor"]
            for col, header in enumerate(summary_headers, 1):
                cell = ws.cell(row=row, column=col)
                cell.value = header
                cell.font = subheader_font
                cell.fill = subheader_fill
                cell.border = border_all
                cell.alignment = Alignment(horizontal="center")
            row += 1

            unique_users = set()
            update_counts = {}
            for item in sorted_updates:
                user_id = item['user_id']
                update_date = item['update']['report_date']
                unique_users.add(user_id)
                if update_date not in update_counts:
                    update_counts[update_date] = 0
                update_counts[update_date] += 1

            summary_data = [
                ["Período do relatório", f"{data_inicial} a {data_final}"],
                ["Total de atualizações", len(sorted_updates)],
                ["Total de usuários", len(unique_users)],
                ["Média de atualizações por usuário", f"{len(sorted_updates)/len(unique_users):.2f}" if unique_users else "0"]
            ]

            for item in summary_data:
                ws.cell(row=row, column=1, value=item[0]).border = border_all
                ws.cell(row=row, column=2, value=item[1]).border = border_all
                row += 1

            logger.debug(f"[DEBUG] Resumo do relatório adicionado à planilha")
        except Exception as e:
            logger.error(f"[DEBUG] Erro ao formatar planilha: {str(e)}")

        file_name = f"relatorio_daily_{data_inicial}_{data_final}.xlsx"

        try:
            logger.debug(f"[DEBUG] Salvando planilha em {file_name}")
            wb.save(file_name)
            logger.debug(f"[DEBUG] Planilha salva com sucesso")
        except Exception as e:
            logger.error(f"[DEBUG] Erro ao salvar planilha: {str(e)}")
            await interaction.followup.send(
                content=f"❌ Erro ao gerar o arquivo Excel: {str(e)}",
                ephemeral=True
            )
            log_command("ERRO", interaction.user, f"/relatorio-daily data_inicial={data_inicial} data_final={data_final}",
                       f"Erro ao salvar arquivo Excel: {str(e)}")
            return

        try:
            logger.debug(f"[DEBUG] Enviando arquivo {file_name} para o Discord")
            await interaction.followup.send(
                content=f"📊 Relatório de atualizações diárias ({data_inicial} a {data_final})",
                file=discord.File(file_name),
                ephemeral=True
            )
            logger.debug(f"[DEBUG] Arquivo enviado com sucesso")

            log_command("RELATÓRIO", interaction.user, f"/relatorio-daily data_inicial={data_inicial} data_final={data_final}",
                       f"Relatório Excel gerado com sucesso")

        except Exception as e:
            logger.error(f"[DEBUG] Erro ao enviar arquivo: {str(e)}")
            await interaction.followup.send(
                content=f"❌ Erro ao enviar o arquivo: {str(e)}",
                ephemeral=True
            )
            log_command("ERRO", interaction.user, f"/relatorio-daily data_inicial={data_inicial} data_final={data_final}",
                       f"Erro ao enviar arquivo: {str(e)}")

        finally:
            try:
                if os.path.exists(file_name):
                    logger.debug(f"[DEBUG] Removendo arquivo temporário {file_name}")
                    os.remove(file_name)
            except Exception as e:
                logger.error(f"[DEBUG] Erro ao remover arquivo temporário: {str(e)}")


class SupportModal(ui.Modal, title="Suporte - Enviar Mensagem"):
    """Modal para envio de mensagens de suporte, erros ou sugestões."""

    support_title = ui.TextInput(
        label="Título",
        style=discord.TextStyle.short,
        placeholder="Ex: Erro ao registrar daily, Sugestão de funcionalidade...",
        required=True,
        min_length=5,
        max_length=100
    )

    support_content = ui.TextInput(
        label="Descrição",
        style=discord.TextStyle.paragraph,
        placeholder="Descreva em detalhes o problema, erro ou sugestão...",
        required=True,
        min_length=10,
        max_length=1000
    )

    async def on_submit(self, interaction: discord.Interaction):
        """Chamado quando o usuário envia o formulário."""
        user_id = str(interaction.user.id)
        title = self.support_title.value
        content = self.support_content.value

        support_user_id = get_env("SUPPORT_USER_ID")

        if not support_user_id:
            await interaction.response.send_message(
                "⚠️ O administrador do sistema não configurou um usuário de suporte. Por favor, entre em contato por outros meios.",
                ephemeral=True
            )
            log_command("ERRO", interaction.user, "/suporte", "SUPPORT_USER_ID não configurado")
            return

        try:
            support_user = await interaction.client.fetch_user(int(support_user_id))

            embed = discord.Embed(
                title=f"📩 Nova Mensagem de Suporte: {title}",
                description=f"**Enviado por:** {interaction.user.mention} ({interaction.user.name}, ID: {interaction.user.id})",
                color=discord.Color.blue()
            )

            embed.add_field(
                name="📝 Conteúdo",
                value=content,
                inline=False
            )

            if interaction.guild:
                embed.add_field(
                    name="🏠 Servidor",
                    value=f"{interaction.guild.name} (ID: {interaction.guild.id})",
                    inline=True
                )

            if interaction.channel:
                embed.add_field(
                    name="📢 Canal",
                    value=f"#{interaction.channel.name} (ID: {interaction.channel.id})",
                    inline=True
                )

            current_time_br = get_br_time().strftime("%d/%m/%Y %H:%M:%S")
            embed.set_footer(text=f"Horário de Brasília: {current_time_br}")

            await support_user.send(embed=embed)

            await interaction.response.send_message(
                "✅ Sua mensagem foi enviada com sucesso para o suporte! Obrigado pelo feedback.",
                ephemeral=True
            )

            log_command("SUPORTE", interaction.user, "/suporte", f"Mensagem enviada: {title}")

        except discord.NotFound:
            await interaction.response.send_message(
                "⚠️ Não foi possível encontrar o usuário de suporte. Por favor, informe o administrador do sistema.",
                ephemeral=True
            )
            log_command("ERRO", interaction.user, "/suporte", f"Usuário de suporte não encontrado: {support_user_id}")
        except discord.Forbidden:
            await interaction.response.send_message(
                "⚠️ O bot não tem permissão para enviar mensagens diretas ao usuário de suporte.",
                ephemeral=True
            )
            log_command("ERRO", interaction.user, "/suporte", "Permissão negada para enviar DM ao suporte")
        except Exception as e:
            await interaction.response.send_message(
                f"❌ Ocorreu um erro ao enviar a mensagem: {str(e)}",
                ephemeral=True
            )
            log_command("ERRO", interaction.user, "/suporte", f"Erro ao enviar mensagem: {str(e)}")


class SupportCommands(commands.Cog):
    """Comandos relacionados ao suporte e feedback."""

    def __init__(self, bot: commands.Bot):
        self.bot = bot

    @app_commands.command(name="suporte", description="Envia uma mensagem de suporte, reporte de erro ou sugestão")
    async def support(self, interaction: discord.Interaction):
        """
        Abre um modal para envio de mensagem de suporte, erro ou sugestão.

        Args:
            interaction: A interação do Discord.
        """
        log_command("COMANDO", interaction.user, "/suporte", "Iniciando modal de suporte")
        await interaction.response.send_modal(SupportModal())


class ConfigView(ui.View):
    """View com botões para escolher configurações do bot."""

    def __init__(self, bot: commands.Bot, funcionalidade: str):
        super().__init__(timeout=300)
        self.bot = bot
        self.funcionalidade = funcionalidade

    @ui.button(label="Configurar Datas Ignoradas", style=discord.ButtonStyle.primary, emoji="📅")
    async def ignored_dates_button(self, interaction: discord.Interaction, button: ui.Button):
        """Botão para configurar datas ignoradas na cobrança de daily."""
        try:
            modal = DateConfigModal(self.bot)
            await interaction.response.send_modal(modal)
            log_command("INFO", interaction.user, f"/config funcionalidade={self.funcionalidade}", "Modal de configuração de datas ignoradas aberto")
        except Exception as e:
            logger = logging.getLogger('team_analysis_bot')
            logger.error(f"Erro ao abrir modal de configuração: {str(e)}")
            await interaction.response.send_message(
                f"❌ Ocorreu um erro ao abrir o modal de configuração: {str(e)}",
                ephemeral=True
            )
            log_command("ERRO", interaction.user, f"/config funcionalidade={self.funcionalidade}", f"Erro ao abrir modal: {str(e)}")

    @ui.button(label="Listar Datas Ignoradas", style=discord.ButtonStyle.secondary, emoji="📋")
    async def list_ignored_dates_button(self, interaction: discord.Interaction, button: ui.Button):
        """Botão para listar as datas ignoradas configuradas."""
        ignored_dates = get_all_ignored_dates()

        if not ignored_dates:
            await interaction.response.send_message(
                "📅 Não há datas configuradas para serem ignoradas na cobrança de daily.",
                ephemeral=True
            )
            log_command("INFO", interaction.user, "/config listar-datas", "Nenhuma data configurada")
            return

        embed = discord.Embed(
            title="📅 Datas Ignoradas - Cobrança de Daily",
            description="Estas são as datas configuradas para serem ignoradas na cobrança de daily:",
            color=discord.Color.blue()
        )

        for date_entry in ignored_dates:
            start_date = datetime.strptime(date_entry["start_date"], "%Y-%m-%d")
            end_date = datetime.strptime(date_entry["end_date"], "%Y-%m-%d")
            created_at = datetime.strptime(date_entry["created_at"], "%Y-%m-%d %H:%M:%S")

            start_date_str = start_date.strftime("%d/%m/%Y")
            end_date_str = end_date.strftime("%d/%m/%Y")
            created_at_str = created_at.strftime("%d/%m/%Y %H:%M:%S")

            if start_date == end_date:
                date_desc = f"📆 **{start_date_str}**"
            else:
                date_desc = f"📆 De **{start_date_str}** até **{end_date_str}**"

            try:
                creator_user = await self.bot.fetch_user(int(date_entry["created_by"]))
                creator_name = creator_user.display_name
            except:
                creator_name = f"Usuário {date_entry['created_by']}"

            embed.add_field(
                name=f"ID: {date_entry['id']} - {date_desc}",
                value=f"Configurado por: {creator_name} em {created_at_str}",
                inline=False
            )

        embed.set_footer(text=f"Total: {len(ignored_dates)} configurações • ID pode ser usado com /remover-data-ignorada")

        await interaction.response.send_message(embed=embed, ephemeral=True)
        log_command("CONSULTA", interaction.user, "/config listar-datas", f"Listadas {len(ignored_dates)} configurações")


async def setup(bot: commands.Bot):
    """Configura todos os cogs de comandos."""
    await bot.add_cog(AdminCommands(bot))
    await bot.add_cog(TimeTrackingCommands(bot))
    await bot.add_cog(UserCommands(bot))
    await bot.add_cog(DailyCommands(bot))
    await bot.add_cog(SupportCommands(bot))